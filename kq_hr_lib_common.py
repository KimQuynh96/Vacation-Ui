import time,datetime, json, random,os,sys,re,codecs
from os import pipe ,path
from openpyxl import load_workbook
from datetime import date
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from selenium import webdriver
from colorama import Fore, Back, Style
from colorama import init, AnsiToWin32
from sys import platform
from pathlib import Path


# Color #
init(wrap = False)
stream = AnsiToWin32(sys.stderr).stream

# Json #
json_file = os.path.dirname(Path(__file__).absolute()) +"\\kq_data_hr.json"

# Excel #
access_page = "Access Page"
functions = "Functions"
now = datetime.datetime.now()
date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
date_id = date_time.replace("/","").replace(", ","").replace(":", "")[2:]
XlsxName = "MenuVacation_"+ str(date_id) +".xlsx"
date_xls = str( datetime.date.today())+ ","+ str(datetime.datetime.now().time())[None:str(datetime.datetime.now().time()).rfind(".")]

# log #
execution_log = os.path.dirname(Path(__file__).absolute()) +"\\Log\\execution_log_" + str(date_id) + ".txt"


if platform == "linux" or platform == "linux2":
    log_folder = "/Log/"
    json_file =  json_file.replace("\\","/")
    execution_log = execution_log.replace("\\","/")
    fail_log = execution_log.replace("execution_log_", "fail_log_")
    error_log = execution_log.replace("execution_log_", "error_log_")
    xlsx_xpath = os.path.dirname(Path(__file__).absolute()) + log_folder + XlsxName
    driver = webdriver.Chrome("/usr/bin/chromedriver")
    driver.maximize_window()
   
else :
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--ignore-ssl-errors=yes")
    driver_path = os.path.dirname(Path(__file__).absolute())+"\\chromedriver.exe"
    driver = webdriver.Chrome(driver_path,chrome_options=chrome_options)
    driver.maximize_window()
    log_folder = "\\Log\\"
    fail_log = execution_log.replace("execution_log_", "fail_log_")
    error_log = execution_log.replace("execution_log_", "error_log_")
    xlsx_xpath = os.path.dirname(Path(__file__).absolute())+log_folder + XlsxName

# json #   
with open(json_file) as json_file:
    data = json.load(json_file)

# create log file of fail test case
open(execution_log, "x").close()

# create log file of fail test case
open(fail_log, "x").close()

# create log file of fail test case
open(error_log, "x").close()

# create .xls 
sheet_name = [functions,access_page]
title = {"1":"Menu","2":"Sub Menu","3":"Test Case Name","4":"Status","5":"Description","6":"Date","7":"Tester"}
wb = openpyxl.Workbook()
wb.save(xlsx_xpath)  
wb1 = load_workbook(xlsx_xpath)
for name in sheet_name :
    wb1.create_sheet(name)
    ws1 = wb1.get_sheet_by_name(name)
    for i in range(1,8):
        ws1.cell(row=1,column=i).value=title[str(i)]
    col = ws1.max_column
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['E'].width = 60
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 15
    my_red  = openpyxl.styles.colors.Color(rgb='00103667')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    for col in range(1,col+1):
        ws1.cell(1,col,value=None).alignment = Alignment(horizontal='center')
        ws1.cell(1,col,value=None).font= Font(size=12, color='FFFFFF', bold=True)
        ws1.cell(1,col,value=None).fill=my_fill
    wb1.save(xlsx_xpath)  
sh = wb1.get_sheet_by_name('Sheet')
wb1.remove_sheet(sh)
wb1.save(xlsx_xpath)


class log():
    def Logging(text):
        log_msg =  codecs.open(execution_log, "a" ,"utf-8")
        log_msg.write(str(text) + "\n")
        log_msg.close()

    def ValidateFailResultAndSystem(fail_msg):
        append_fail_result =  codecs.open(fail_log, "a" ,"utf-8")
        append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
        append_fail_result.close()

class xl():
    def add_data_in_excel(content_excel,status,description,sheet):
        wb = openpyxl.load_workbook(xlsx_xpath) 
        if sheet == "ac":
            sheet_use = wb.get_sheet_by_name(access_page)
        else:
            sheet_use = wb.get_sheet_by_name(functions) 
        row = sheet_use.max_row
        col = sheet_use.max_column
        status = status.replace(" ", "")
        if len(status) != 0:
            if status != "Pass":
                sheet_use.cell(row+1,col-6).font= Font(color='FF0000')
                sheet_use.cell(row+1,col-5).font= Font(color='FF0000')
                sheet_use.cell(row+1,col-4).font= Font(color='FF0000')
                sheet_use.cell(row+1,col-3).font= Font(color='FF0000')
                sheet_use.cell(row+1,col-2).font= Font(color='FF0000')
                sheet_use.cell(row+1,col-1).font= Font(color='FF0000')
                sheet_use.cell(row+1,col).font= Font(color='FF0000')

            content_excel["description"] = description
            sheet_use.cell(row=row+1,column=col-6).value=content_excel["menu"]
            sheet_use.cell(row=row+1,column=col-5).value=content_excel["submenu"]
            sheet_use.cell(row=row+1,column=col-4).value=content_excel["testcase"]
            sheet_use.cell(row=row+1,column=col-3).value=content_excel["status"]
            sheet_use.cell(row=row+1,column=col-2).value=content_excel["description"]
            sheet_use.cell(row=row+1,column=col-1).value=date_xls
            sheet_use.cell(row=row+1,column=col).value=content_excel["tester"]
            wb.save(xlsx_xpath)
        else:
            sheet_use.cell(row=row+1,column=col-6).value=content_excel["menu"]
            sheet_use.cell(row=row+1,column=col-5).value=content_excel["submenu"]
            sheet_use.cell(row=row+1,column=col-4).value=content_excel["testcase"]
            sheet_use.merge_cells("D" + str(row+1) + ":G" + str(row+1))
            sheet_use.cell(row=row+1,column=4).value = description
            wb.save(xlsx_xpath)

class cm():
    '''
    def xlsx(content_excel,text):
        if content_excel["status"] == "Pass":
            log.Logging(text)
            print("\033[32m"+text+"\033[39m")
        else :
            log.ValidateFailResultAndSystem(text)
            print("\033[31m"+text+"\033[39m")
        description = text[None:int(text.rfind("<"))]
        description = description.lstrip()[1:None]
        status = content_excel["status"]
        sheet = content_excel["sheet"]
        xl.add_data_in_excel(content_excel,status,description,sheet)
    '''
    def xlsx(content_excel,text):
       
        if content_excel["status"] == "Pass":
            log.Logging(text)
            print("\033[32m"+text+"\033[39m")
        else :
            log.ValidateFailResultAndSystem(text)
            print("\033[31m"+text+"\033[39m")
        description = text[None:int(text.rfind("<"))]
        description = description.lstrip()[1:None]
        status = content_excel["status"]
        sheet = content_excel["sheet"]
        xl.add_data_in_excel(content_excel,status,description,sheet)



    def msg(t,text):
        if t == "p":
            log.Logging(text)
            print("\033[32m"+text+"\033[39m")
        elif t == "n":
            log.Logging(text)
            print("\033[33m"+text+"\033[39m")
        elif t == "t":
            log.Logging(text)
            print("\033[37m"+text+"\033[39m")
        else :
            log.ValidateFailResultAndSystem(text)
            print("\033[31m"+text+"\033[39m")
   
   
    def excel(result_excel,status,testcase):
        #result_excel["Select vacation date"]="Pass"
        if status == "p":
            result_excel[testcase] = "Pass"
        else:
            result_excel[testcase] = "Fail"


    def login_result():
        try :
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID,"menubar")))
            return True
        except NoSuchElementException:
            return False
        
    def select_user(user_name):
        ip_search_user = driver.find_element_by_xpath(data["admin"]["ip_search_user"])
        ip_search_user.send_keys(user_name)
        ip_search_user.send_keys(Keys.RETURN)
        time.sleep(2)
        driver.find_element_by_css_selector(data["admin"]["select_user"]).click()
        driver.find_element_by_xpath(data["admin"]["bt_add_user"]).click()
        driver.find_element_by_xpath(data["admin"]["bt_save_user"]).click()

    def popup_time_card():
        try :
            driver.find_element_by_css_selector(".pb-1 > .feather").click()
        except :
            pass

    def is_Displayed(xpath):
        try:
            time.sleep(1)
            driver.find_element_by_xpath(xpath)
            return True
        except NoSuchElementException:
            return False

    def is_Displayed1(typexpath,xpath):
        if typexpath =="textlink":
            try:
                driver.find_element_by_link_text(xpath)
                return True
            except NoSuchElementException:
                return False
        if typexpath =="id":
            try:
                WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID,xpath)))
                return True
            except NoSuchElementException:
                return False

    def scroll():
        html = driver.find_element_by_tag_name('html')
        html.send_keys(Keys.END)
    
    def scrolling_to_target(target):
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_to_element(target).perform()
    
    def total_data(list):
        total = 0
        for element in list :
            total = total + 1
        return total

    def xpath(element,i,xpath):
        
        return "//"+element+"["+str(i)+ "]" + xpath

    def xpath1(element,xpath):
        return "//"+element + xpath

    def xpath2(element,i,xpath):
    
        return element + str(i) +  xpath

    def xpath3(element,i,xpath,j,xpath1):
    
        return "//"+element+"["+str(i)+ xpath + str(j)+xpath1

    def organization(from_element,to_element):
        touch = ActionChains(driver)
        touch.click_and_hold(from_element).move_to_element(to_element).release(to_element)
           
    def alert():
        WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.ID, "noty_layout__topRight")))
        alert = {"result":"","content":""}
        notification = driver.execute_script('return document.getElementById("noty_layout__topRight").innerText')
        content_notification = notification.split("\n")
        alert["result"] = content_notification[0]
        alert["content"] = content_notification[1]
        print("alert :",alert)
        return alert

    def param_data():
        today = date.today()
        current_month = today.month
        current_year = today.year
        date_time = date.today() 
        par =  {
            "current_month":today.month,
            "month":data["month"],
            "time":"//div[contains(text(),'" +str(current_year)+".0"+str(current_month)+"')]",
            "vacation_name":"Vacation " + str(random.randint(0,10000)) + "[" +str(date_time) + "]",
            "number_day_off":12,
        }
        return json.dumps(par)

    def param_url(domain):
        par = {
            "url_login":"http://"+domain+"/ngw/app/#/sign",
            "menu_vacation":"http://"+domain+"/ngw/app/#/nhr",
            "vacation_status":"http://"+domain+"/php7/rain/laravel/hr/holiday/user/vacation/status/request",
            "request_vacation":"http://"+domain+"/php7/rain/laravel/hr/holiday/user/app",
            "create_vacation":"http://"+domain+"/php7/rain/laravel/hr/holiday/admin/holiday/vacations",
            "add_manager":"http://"+domain+"/php7/rain/laravel/hr/holiday/admin/manager/manager",
            "arbitrary_decision_setting":"http://"+domain+"/php7/rain/laravel/hr/holiday/admin/manager/approval",
            "basic_settings":"http://"+domain+"/php7/rain/laravel/hr/holiday/admin/basic",
        }

        return json.dumps(par)

        

    